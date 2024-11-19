import openpyxl
from django.core.management.base import BaseCommand
from openpyxl.reader.excel import load_workbook
# Странно но факт тут должна быть ошибка(от компилятора)
from invAPP.models import Workers, Inventory1C, Characteristics, Relation3, Relation4, TypeOfEquipment


class Command(BaseCommand):
    help = 'Импортирует данные из Excel в базы данных Django'

    def add_arguments(self, parser):
        parser.add_argument('excel_files', type=str)

    def handle(self, *args, **options):
        excel_files = options['excel_files']

        # for file in excel_files:
        self.import_data(excel_files)

    def import_data(self, file):
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            self.stdout.write(self.style.SUCCESS(f"Обрабатываем файл: {file}"))
            headers = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

            self.stdout.write(f"Импорт в таблицу <Type of equipment>")
            types_of_equipment = [
                "Иное", "ПК", "Моноблок", "Ноутбук", "Сервер", "Планшет",
                "КПК, смартфон", "Моб.телефон", "Факс", "ИБП", "Монитор",
                "Телевизор", "Видеомагнитофон", "Видеокамера аналог.", "Звуковое оборудование", "Микшерный пульт",
                "Проектор", "Интерактивная доска", "Интерактивная панель",
                "Принтер", "МФУ", "Копир", "Сканер", "Документ-проектор",
                "АТС", "Видеорегистратор", "СКУД", "Коммутатор, сетевое"
            ]
            for type_name in types_of_equipment:
                obj, created = TypeOfEquipment.objects.get_or_create(type_name=type_name)
            self.stdout.write("Импорт данных завершён в <Type of equipment>.")

            # эти данные можно извлечь из 2 таблиц но в одной есть перво начальная стоимость в бдругой нет
            if ("ФИО" in headers and "Подразделение" in headers) or ("Инвентарный номер" in headers and ("ОС" in headers or "Основное средство" in headers )and "Дата принятия к учету" in headers):
                if ("ФИО" in headers and "Подразделение" in headers):
                    fio_col = headers["ФИО"]
                    department_col = headers["Подразделение"]
                self.stdout.write(f"Импорт в таблицу <Workers> and <Inventory1C>")
                if ("Инвентарный номер" in headers and ("ОС" in headers or "Основное средство" in headers )and "Дата принятия к учету" in headers):
                    inv_col = headers["Инвентарный номер"]
                    if "ОС" in headers:
                        gen_col = headers["ОС"]
                    elif "Основное средство" in headers:
                        gen_col = headers["Основное средство"]
                    date_col = headers["Дата принятия к учету"]
                    if "Стоимость первоначальна" in headers:
                        init_col = headers["Стоимость первоначальна"]
                    else:
                        init_col = headers["Балансовая стоимость"]  # ПОка заменил на балансовую из 1 таблицы(где цифры)

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    #все в одном цикле т.к. мне нужно получить id_workers(это приминимо к 1 таблице а не ко второй(1С)б там я хз как)
                    if ("ФИО" in headers and "Подразделение" in headers):
                        fio_value = row[fio_col - 1]
                        department_value = row[department_col - 1]
                        if fio_value and department_value:
                            worker, created = Workers.objects.get_or_create(
                                full_name=fio_value,
                                department = department_value,
                            )
                            id_worker = worker.id_workers

                    if ("Инвентарный номер" in headers and (
                            "ОС" in headers or "Основное средство" in headers) and "Дата принятия к учету" in headers):
                        inv_value = row[inv_col - 1]
                        gen_value = row[gen_col - 1]
                        date_value = row[date_col - 1]
                        init_value = row[init_col - 1]

                        if inv_value and init_value and date_value and gen_value:
                            Inventory1C.objects.get_or_create(
                            accounting_name = gen_value,
                            inventory_decimal=inv_value,
                            date_acceptance_accounting=date_value,
                            defaults={
                                "real_name": "", #реальное имя я хз где получить
                                "date_of_decommission": ""
                            },
                            initial_cost = init_value,
                            id_workers = id_worker, #так же не понятно откуда взять, если из 1С брать данные
                            id_type = TypeOfEquipment.objects.get(type_name=type_name).id_type,

                            )
                self.stdout.write("Импорт данных завершён в <Workers> and <Inventory1C>.")
        except Exception as e:
            self.stderr.write(f"Ошибка при импорте данных: {e}")