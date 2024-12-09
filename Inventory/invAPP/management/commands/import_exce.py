import openpyxl
from django.core.management.base import BaseCommand
from openpyxl.reader.excel import load_workbook
# Странно но факт тут должна быть ошибка(от компилятора)
from invAPP.models import Workers, Inventory1C, Characteristics, Relation3, Relation4, TypeOfEquipment

from datetime import datetime


class Command(BaseCommand):
    help = 'Импортирует данные из двух Excel-файлов в базы данных Django'

    def add_arguments(self, parser):
        parser.add_argument('excel_files', nargs=2, type=str, help="Пути к двум файлам Excel.")

    def handle(self, *args, **options):
        file1, file2 = options['excel_files']
        try:
            data1 = self.load_excel(file1)  # 1C
            data2 = self.load_excel(file2)  # 101.34 на 31.12.22 v2

            self.stdout.write(f"Обработка данных из файлов {file1} и {file2}")

            types_of_char = [
                "иное", "ПК", "моноблок", "Ноутбук", "Сервер", "Планшет",
                "КПК, смартфон", "Моб.телефон", "Факс", "ИБП", "Монитор",
                "Телевизор", "Видеомагнитофон", "Видеокамера аналог.", "Звуковое оборудование", "Микшерный пульт",
                "Проектор", "Интерактивная доска", "Интерактивная панель",
                "Принтер", "МФУ", "Копир", "Сканер", "Документ-проектор",
                "АТС", "Видеорегистратор", "СКУД", "Коммутатор, сетевое"
            ]
            types = [
                "отечественный", "общедоступный", "неизвестно"
            ]
            type_equipment = [
                "ПК отечественный".lower(),
                "Ноутбук отечественный".lower(),
                "Моноблок отечественный".lower(),
                "Общедоступный ПК".lower()
            ]
            characteristcis = [char_name.lower() for char_name in types_of_char]
            for char_name in characteristcis:
                Characteristics.objects.get_or_create(char_name=char_name)

            for type_name in types:
                TypeOfEquipment.objects.get_or_create(
                    type_name=type_name,
                )
            type_equipment_map = {
                "ПК отечественный".lower(): "отечественный",
                "Ноутбук отечественный".lower(): "отечественный",
                "Моноблок отечественный".lower(): "отечественный",
                "Общедоступный ПК".lower(): "общедоступный"
            }
            type_mapping_char = {
                "ПК 2021".lower(): "ПК",
                "ПК отечественный".lower(): "ПК".lower(),
                "Ноутбук отечественный".lower(): "Ноутбук".lower(),
                "Моноблок отечественный".lower(): "Моноблок".lower(),
                "Комп 2022".lower(): "ПК".lower(),
                "комп 2021".lower(): "ПК".lower(),
                "Комп. не более 5 лет".lower(): "ПК".lower(),
                "Презен. об. не более 5 лет".lower(): "Проектор".lower(),
                "Пр., мфу, коп., скан. не более 5 лет".lower(): "Принтер".lower(),
            }

            inventory_map = {}
            self.stdout.write("Идем по второму файлу")
            for row in data2:
                inv_number = row.get('Инвентарный номер')
                if inv_number:
                    inventory_map[inv_number] = row
                fio = row.get('ФИО')
                department = row.get('Подразделение')
                if not fio or not department:
                    fio = "Неизвестно"
                    department = "Неизвестно"

                worker, _ = Workers.objects.get_or_create(
                    full_name=fio,
                    department=department
                )

                initial_cost = row.get('Стоимость первоначальна')

                if not initial_cost:
                    initial_cost = 0

                type_eq = None
                type_char = None
                for col, value in row.items():
                    col = col.strip().lower()
                    if col in types_of_char:
                        if str(value).strip() == "1":
                            type_name = type_mapping_char.get(col, col)
                            type_char, _ = Characteristics.objects.get_or_create(char_name=type_name)
                            break
                for col, value in row.items():
                    col = col.strip().lower()
                    if col in type_equipment:
                        if str(value).strip() == "1":
                            type_name_eq = type_equipment_map.get(col, col)
                            type_eq, _ = TypeOfEquipment.objects.get_or_create(type_name=type_name_eq)
                            break
                # self.stdout.write(f"Обработка данных из файлов {fio} и {department} и {type_obj}")
                if type_char is None:
                    type_char, _ = Characteristics.objects.get_or_create(char_name="иное")
                if type_eq is None:
                    type_eq, _ = TypeOfEquipment.objects.get_or_create(type_name="неизвестно")
                date_acceptance = row.get('Дата принятия к учету')
                date_decommission = row.get('Дата списания')
                if not date_decommission:
                    date_decommission = ""
                if not row.get('ОС'):
                    continue

                inventory_obj, _ = Inventory1C.objects.get_or_create(
                    inventory_decimal=inv_number,
                    defaults={
                        'id_workers': worker,
                        'accounting_name': row.get('ОС'),
                        'date_acceptance_accounting': date_acceptance,
                        'initial_cost': initial_cost,
                        'id_type': type_eq,
                        "real_name": "",
                        "date_of_decommission": date_decommission
                    }

                )
                Relation3.objects.get_or_create(
                    type_of_equipment_id_type=type_eq,
                    characteristics_id_char=type_char,
                )
                Relation4.objects.get_or_create(
                    inventory_1c_id=inventory_obj,
                    characteristics_id_char=type_char,
                    defaults={
                        "value": ""
                    }
                )

            self.stdout.write("Идем по первому файлу")
            for row in data1:
                inv_number = row.get('Инвентарный номер')
                if not inv_number:
                    inv_number = "Неизвестно"
                fio = "Неизвестно"
                department = "Неизвестно"

                worker, _ = Workers.objects.get_or_create(
                    full_name=fio,
                    department=department
                )
                initial_cost = row.get('Стоимость первоначальна')
                if not initial_cost:
                    initial_cost = 0
                type_eq = None
                type_char = None

                for type_name in types_of_char:
                    if type_name in row and str(row[type_name]).strip():
                        type_char, _ = Characteristics.objects.get_or_create(char_name=type_name)
                        self.stdout.write(f"Идем по первому файлу {type_char}")
                        break
                for type_name_eq in types:
                    if type_name_eq in row and str(row[type_name_eq]).strip():
                        type_eq, _ = TypeOfEquipment.objects.get_or_create(type_name=type_name_eq)
                        self.stdout.write(f"Идем по первому файлу {type_eq}")
                        break
                if type_char is None:
                    type_char, _ = Characteristics.objects.get_or_create(char_name="иное")
                if type_eq is None:
                    type_eq, _ = TypeOfEquipment.objects.get_or_create(type_name="неизвестно")

                date_acceptance = row.get('Дата принятия к учету')
                if not row.get('Основное средство'):
                    continue
                inventory_obj, _ = Inventory1C.objects.get_or_create(
                    inventory_decimal=inv_number,
                    id_workers=worker,
                    accounting_name=row.get('Основное средство'),
                    date_acceptance_accounting=date_acceptance,
                    initial_cost=initial_cost,
                    id_type=type_eq,
                    defaults={
                        "real_name": "",
                        "date_of_decommission": ""
                    }
                )

                Relation3.objects.get_or_create(
                    type_of_equipment_id_type=type_eq,
                    characteristics_id_char=type_char,
                )
                Relation4.objects.get_or_create(
                    inventory_1c_id=inventory_obj,
                    characteristics_id_char=type_char,
                    defaults={
                        "value": ""
                    }

                )

            self.stdout.write(self.style.SUCCESS("Импорт данных завершён."))

        except Exception as e:
            self.stderr.write(f"Ошибка: {e}")

    def load_excel(self, file):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        data = [
            {headers[idx]: cell for idx, cell in enumerate(row)}
            for row in sheet.iter_rows(min_row=2, values_only=True)
        ]
        return data

    def get_merged_cell_value(self, row, column_name):
        column_idx = list(row.keys()).index(column_name)
        cell_value = row.get(column_name)

        if cell_value is None:
            return None

        return cell_value

# Пока дату преобразовал в string так что метод не используется
def convert_date(date_string):
    if date_string:
        try:
            return datetime.strptime(date_string.strip(), '%d.%m.%Y').date()
        except ValueError:
            print(f"Ошибка преобразования даты: {date_string}")
            return None
    return None